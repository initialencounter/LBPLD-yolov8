import os
from openpyxl import load_workbook
from PIL import Image as PILImage

def extract_images_from_xlsx(xlsx_path, output_dir):
    try:
        workbook = load_workbook(xlsx_path)
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            for image in sheet._images:
                img_filename = f"{image.__hash__()}.png"
                img_path = os.path.join(output_dir, img_filename)
                PILImage.open(image.ref).convert('RGB').save(img_path)
    except Exception as e:
        print(f"Error processing {xlsx_path}: {e}")

def recursive_traversal_and_extract_images(root_dir, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for subdir, _, files in os.walk(root_dir):
        for file in files:
            file_path = os.path.join(subdir, file)
            if file.endswith('.xlsx'):
                extract_images_from_xlsx(file_path, output_dir)

if __name__ == "__main__":
    root_directory = "E:\\2024\\5"
    output_directory = "./datasets/image"
    recursive_traversal_and_extract_images(root_directory, output_directory)

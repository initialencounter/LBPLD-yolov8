import os
import time
from win32com.client import Dispatch
from PIL import ImageGrab

from openpyxl import load_workbook
from PIL import Image as PILImage

def extract_images_from_xls(excel_path, output_folder):
    try:
        # 创建Excel应用程序对象
        excel = Dispatch("Excel.Application")
        # 打开Excel文件
        workbook = excel.Workbooks.Open(excel_path)
        # 遍历所有工作表
        for sheet in workbook.Sheets:
            # 遍历每个工作表中的每个Shape对象
            for shape in sheet.Shapes:
                if shape.Type == 13:  # msoPicture类型
                    # 使用ImageGrab从剪贴板中获取图片
                    shape.Copy()
                    image = ImageGrab.grabclipboard()
                    if image:

                        
                        save_image(image, output_folder)
        # 关闭工作簿
        workbook.Close(SaveChanges=False)
        # 退出Excel应用程序
        excel.Quit()
    except Exception as e:
        print(f"Error processing: {e}")

def extract_images_from_xlsx(xlsx_path, output_folder):
    try:
        workbook = load_workbook(xlsx_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for image in sheet._images:
                save_image(PILImage.open(image.ref).convert('RGB'), output_folder)
    except Exception as e:
        print(f"Error processing {xlsx_path}: {e}")
        
def save_image(image, output_folder):
    if image.size[0] * image.size[1] > 65792:
        img_filename = f"{time.time()}.png"
        image_path = os.path.join(output_folder, img_filename)
        image.save(image_path)
        print(image.size[0] * image.size[1], img_filename)

# 使用示例
if __name__ == "__main__":
    extract_images_from_xls(os.path.abspath(r'../xls/1.装箱资料LR-FK202.xls'), './')
